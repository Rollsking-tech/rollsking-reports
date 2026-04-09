import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="RollsKing Reports", page_icon="🍱",
                   layout="centered", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: #0f0f0f; color: #f0f0f0; }
.main-title { font-family:'Syne',sans-serif; font-size:2.4rem; font-weight:800; color:#fff; letter-spacing:-1px; line-height:1.1; margin-bottom:0.2rem; }
.main-subtitle { color:#666; font-size:0.95rem; margin-bottom:1.5rem; }
.section-label { font-family:'Syne',sans-serif; font-size:0.78rem; font-weight:700; color:#e8a020; letter-spacing:1.5px; text-transform:uppercase; margin-bottom:0.4rem; }
.card { background:#1a1a1a; border:1px solid #2a2a2a; border-radius:12px; padding:1.1rem 1.3rem; margin-bottom:1rem; }
.status-ok { color:#4ade80; font-weight:600; font-size:0.9rem; }
.chip-ok   { background:#14532d; color:#4ade80; border-radius:6px; padding:3px 10px; font-size:0.82rem; display:inline-block; margin:2px 0; }
.chip-warn { background:#7c2d12; color:#fca5a5; border-radius:6px; padding:3px 10px; font-size:0.82rem; display:inline-block; margin:2px 0; }
.step-badge { background:#e8a020; color:#000; border-radius:50%; width:22px; height:22px; display:inline-block; text-align:center; line-height:22px; font-weight:800; font-size:0.8rem; margin-right:8px; }
hr.divider { border:none; border-top:1px solid #2a2a2a; margin:1.5rem 0; }
.stButton > button { background:#e8a020 !important; color:#000 !important; font-family:'Syne',sans-serif !important; font-weight:700 !important; border-radius:8px !important; border:none !important; padding:0.55rem 1.5rem !important; font-size:0.9rem !important; }
.stButton > button:hover { background:#f5b535 !important; }
div[data-testid="stExpander"] { background:#1a1a1a !important; border:1px solid #2a2a2a !important; border-radius:10px !important; }
.stTabs [data-baseweb="tab-list"] { background:#1a1a1a !important; border-radius:10px !important; }
.stTabs [aria-selected="true"] { color:#e8a020 !important; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def safe_id(v):
    try:
        s = str(v).strip()
        return None if s in ('#N/A','','None','nan','RF NOT AVL','MISSING — fill in') else int(float(s))
    except:
        return None

def safe_f(v, d=0.0):
    try:
        f = float(v if v is not None else d)
        return d if f != f else f
    except:
        return d

def parse_pct(v):
    if v is None: return None
    try:
        s = str(v).replace('%','').strip()
        f = float(s)
        return f if f > 1 else f * 100
    except:
        return None

def parse_min(v):
    if v is None: return None
    try:
        return float(str(v).replace(' mins','').replace(' min','').strip())
    except:
        return None

# ── SCORING RULES (hardcoded from confirmed master sheet) ─────────────────────
def score_complaints(pct):
    """0-1% = 4pts | 1-2% = 3pts | 2-3% = 1pt | >=3% = 0pt"""
    if pct is None: return 0
    if pct < 1:     return 4
    if pct < 2:     return 3
    if pct < 3:     return 1
    return 0

def score_fc(pct):
    """< 40% = 1pt | >= 40% = 0pt"""
    if pct is None: return 0
    return 1 if pct < 40 else 0

def score_kpt(avg_min):
    """floor(avg) <= 12 = 1pt | floor(avg) > 12 = 0pt
    Confirmed: manual applies floor() then <= 12 threshold"""
    if avg_min is None: return 0
    import math
    return 1 if math.floor(avg_min) <= 12 else 0

def score_rating(avg):
    """>= 4.0 = 1pt | < 4.0 = 0pt"""
    if avg is None or avg == 0: return 0
    return 1 if avg >= 4.0 else 0

def score_avail(pct):
    """>= 98% = 1pt | < 98% = 0pt"""
    if pct is None: return 0
    return 1 if pct >= 98 else 0

def score_tier(avg):
    """Bronze 0-3 | Silver 3-6 | Gold 6-8 | Platinum 8+"""
    if avg >= 8: return "Platinum"
    if avg >= 6: return "Gold"
    if avg >= 3: return "Silver"
    return "Bronze"

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════
TIER_CLR = {
    "Platinum": ("1F1F1F","FFD700"),
    "Gold":     ("1F1F1F","FFA500"),
    "Silver":   ("1F1F1F","C0C0C0"),
    "Bronze":   ("FFFFFF","8B4513"),
}
CLR = {
    "hd":"1F2D3D","hm":"2E4057","wh":"FFFFFF","lg":"F2F2F2",
    "mg":"D9D9D9","gn":"C6EFCE","rd":"FFC7CE","yw":"FFF2CC",
}

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row_num, cols, bg, fg="FFFFFF", sz=9):
    for col_idx, text in enumerate(cols, 1):
        c = ws.cell(row=row_num, column=col_idx, value=text)
        c.font      = Font(bold=True, color=fg, size=sz, name="Arial")
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()

# ══════════════════════════════════════════════════════════════════════════════
# HARDCODED MAPPING  (from confirmed Team_Performance_Dashboard__4_.xlsx)
# IDs confirmed:
#   zmt_rk / zmt_rf = Zomato RollsKing / Rolling Fresh restaurant IDs
#   swg_rk / swg_rf = Swiggy RollsKing / Rolling Fresh restaurant IDs
#   pos             = PetPooja POS ID
#   None on RF IDs  = outlet confirmed as RK-only ("RF NOT AVL")
# ══════════════════════════════════════════════════════════════════════════════
DEFAULT_MAPPING = {
  "Navneet Singh": [
    {"outlet":"Sec 104",             "pos":28039,  "zmt_rk":19476740,"zmt_rf":20884624,"swg_rk":313666, "swg_rf":783919},
    {"outlet":"Sector-141 Noida",    "pos":26592,  "zmt_rk":18734595,"zmt_rf":21198454,"swg_rk":63465,  "swg_rf":882814},
    {"outlet":"Sector-132 Noida",    "pos":34303,  "zmt_rk":18750756,"zmt_rf":21563311,"swg_rk":68184,  "swg_rf":998424},
    {"outlet":"Sector 125 Noida",    "pos":373602, "zmt_rk":21824279,"zmt_rf":21819721,"swg_rk":1069514,"swg_rf":1069500},
    {"outlet":"Sector-73 Noida",     "pos":97074,  "zmt_rk":20508934,"zmt_rf":20511693,"swg_rk":635149, "swg_rf":641252},
    {"outlet":"Sector-44 Noida",     "pos":39966,  "zmt_rk":18575970,"zmt_rf":21570341,"swg_rk":42813,  "swg_rf":1035832},
  ],
  "Ajay Halder": [
    {"outlet":"Sector 4 Noida",      "pos":21787,  "zmt_rk":19364731,"zmt_rf":20884617,"swg_rk":54622,  "swg_rf":784582},
    {"outlet":"Sector-62",           "pos":23687,  "zmt_rk":302308,  "zmt_rf":21292616,"swg_rk":42808,  "swg_rf":907194},
    {"outlet":"Sector-37",           "pos":32851,  "zmt_rk":20374787,"zmt_rf":20884635,"swg_rk":583789, "swg_rf":798506},
    {"outlet":"Sector-18",           "pos":26952,  "zmt_rk":304612,  "zmt_rf":22333370,"swg_rk":42807,  "swg_rf":1238358},  # RF verified
    {"outlet":"Gaur City GNoida",    "pos":112772, "zmt_rk":20589872,"zmt_rf":21087471,"swg_rk":879431, "swg_rf":940381},
    {"outlet":"Eco Loft",            "pos":74178,  "zmt_rk":20264919,"zmt_rf":22245955,"swg_rk":531120, "swg_rf":1227942},
  ],
  "Sunil Sharma": [
    {"outlet":"RDC Raj Nagar Gzb",   "pos":363143, "zmt_rk":18962941,"zmt_rf":21184529,"swg_rk":879460, "swg_rf":883371},
    {"outlet":"GNB Mall",            "pos":113953, "zmt_rk":21341669,"zmt_rf":21340893,"swg_rk":1082869,"swg_rf":1089374},
    {"outlet":"Shipra Mall",         "pos":408910, "zmt_rk":22103426,"zmt_rf":None,    "swg_rk":1238359,"swg_rf":None},    # RK only
  ],
  "Vishwanath Rao": [
    {"outlet":"Indirapuram",         "pos":38041,  "zmt_rk":18633334,"zmt_rf":20884637,"swg_rk":46674,  "swg_rf":784485},
    {"outlet":"Rajendra Nagar Gzb",  "pos":37055,  "zmt_rk":19283683,"zmt_rf":21563308,"swg_rk":241917, "swg_rf":998786},
    {"outlet":"Vasundhra",           "pos":122466, "zmt_rk":20711593,"zmt_rf":21780276,"swg_rk":731841, "swg_rf":1069774},
  ],
  "Sanjay Morya": [
    {"outlet":"Nathupur Gurugram",   "pos":108782, "zmt_rk":20582763,"zmt_rf":20884615,"swg_rk":668475, "swg_rf":783922},
    {"outlet":"Old DLF Sec-14 Gurgaon","pos":108777,"zmt_rk":20582827,"zmt_rf":20884599,"swg_rk":668470,"swg_rf":854941},
    {"outlet":"Sector-57 Gurugram",  "pos":74068,  "zmt_rk":20463325,"zmt_rf":20964530,"swg_rk":624165, "swg_rf":803919},
    {"outlet":"Wazirabad Gurugram",  "pos":108779, "zmt_rk":20582847,"zmt_rf":None,    "swg_rk":668467, "swg_rf":None},    # RK only
    {"outlet":"Gurugram Sec-82",     "pos":30407,  "zmt_rk":19513923,"zmt_rf":21087476,"swg_rk":327106, "swg_rf":850542},
    {"outlet":"Sector 90 Gurugram",  "pos":380769, "zmt_rk":21929020,"zmt_rf":21929049,"swg_rk":1102249,"swg_rf":1101896},
  ],
  "Harinder Nagar": [
    {"outlet":"Kalkaji",             "pos":31247,  "zmt_rk":18869459,"zmt_rf":20884610,"swg_rk":90719,  "swg_rf":783920},
    {"outlet":"Tilak Nagar",         "pos":63819,  "zmt_rk":18942689,"zmt_rf":20884619,"swg_rk":123197, "swg_rf":786729},
    {"outlet":"Vasant Kunj",         "pos":25924,  "zmt_rk":19030978,"zmt_rf":21198457,"swg_rk":131217, "swg_rf":883284},
    {"outlet":"Chattarpur",          "pos":26423,  "zmt_rk":19052007,"zmt_rf":21198467,"swg_rk":140433, "swg_rf":883277},
    {"outlet":"Paschim Vihar",       "pos":43412,  "zmt_rk":20256577,"zmt_rf":21751304,"swg_rk":531480, "swg_rf":1065481},
    {"outlet":"Gtb Nagar",           "pos":79050,  "zmt_rk":20323930,"zmt_rf":21929075,"swg_rk":569414, "swg_rf":1101863},
    {"outlet":"Rohini",              "pos":93493,  "zmt_rk":22100897,"zmt_rf":22101011,"swg_rk":622353, "swg_rf":1167156},
    {"outlet":"Vikashpuri",          "pos":404584, "zmt_rk":22227640,"zmt_rf":22227658,"swg_rk":1224350,"swg_rf":1224348},
    {"outlet":"Uttam Nagar Dwarka",  "pos":414828, "zmt_rk":None,    "zmt_rf":None,   "swg_rk":1280807,"swg_rf":1280814},
    {"outlet":"Subhash Nagar",       "pos":398993, "zmt_rk":22165860,"zmt_rf":22165921,"swg_rk":1196325,"swg_rf":1196321},
  ],
  "Zeeshan Ali": [
    {"outlet":"Shaheen Bagh",        "pos":118685, "zmt_rk":20666436,"zmt_rf":21190578,"swg_rk":704360, "swg_rf":879433},
    {"outlet":"NIT Faridabad",       "pos":96843,  "zmt_rk":20480333,"zmt_rf":21087481,"swg_rk":632083, "swg_rf":852302},
    {"outlet":"Sec-15 Faridabad",    "pos":54369,  "zmt_rk":18567324,"zmt_rf":21702217,"swg_rk":42815,  "swg_rf":1037447},
    {"outlet":"Lakkarpur Faridabad", "pos":143500, "zmt_rk":20873208,"zmt_rf":21087485,"swg_rk":775707, "swg_rf":855113},
    {"outlet":"Greenfield Faridabad","pos":154254, "zmt_rk":21446399,"zmt_rf":21446783,"swg_rk":983943, "swg_rf":991036},
  ],
  "Badir Alam": [
    {"outlet":"Bhopal",              "pos":338959, "zmt_rk":21340655,"zmt_rf":21340565,"swg_rk":934354, "swg_rf":937374},
    {"outlet":"Indore",              "pos":109589, "zmt_rk":20566161,"zmt_rf":21304975,"swg_rk":673809, "swg_rf":920802},
    {"outlet":"Siddharth Nagar Indore","pos":156653,"zmt_rk":21022031,"zmt_rf":21643899,"swg_rk":690867,"swg_rf":1027884},
  ],
  "Abhishek Kumar": [
    {"outlet":"Whitefield Bangalore","pos":89397,  "zmt_rk":20410563,"zmt_rf":21075165,"swg_rk":606509, "swg_rf":850483},
    {"outlet":"Mahadevpura",         "pos":72269,  "zmt_rk":20201048,"zmt_rf":20790266,"swg_rk":515199, "swg_rf":700101},
    {"outlet":"Koramangala",         "pos":83769,  "zmt_rk":20359621,"zmt_rf":20790279,"swg_rk":580691, "swg_rf":709590},
    {"outlet":"Electronic City",     "pos":72413,  "zmt_rk":20213913,"zmt_rf":21087516,"swg_rk":515053, "swg_rf":848482},
    {"outlet":"Sarjapur",            "pos":68691,  "zmt_rk":20163232,"zmt_rf":20790275,"swg_rk":494751, "swg_rf":649361},
    {"outlet":"Kalyan Nagar",        "pos":75899,  "zmt_rk":20265149,"zmt_rf":21087530,"swg_rk":544214, "swg_rf":848481},
    {"outlet":"Bel Road Bangalore",  "pos":75897,  "zmt_rk":20263151,"zmt_rf":21037571,"swg_rk":536015, "swg_rf":728281},
    {"outlet":"Habble Bangalore",    "pos":95682,  "zmt_rk":20471662,"zmt_rf":None,    "swg_rk":625912, "swg_rf":None},    # RK only
    {"outlet":"Indira Nagar Bangalore","pos":403199,"zmt_rk":22179137,"zmt_rf":22179218,"swg_rk":1203098,"swg_rf":1203101},
  ],
  "Virendra Pratap": [
    {"outlet":"Mohanram Nagar",      "pos":84743,  "zmt_rk":20410863,"zmt_rf":20994725,"swg_rk":588878, "swg_rf":808838},
    {"outlet":"Madipakkam",          "pos":84742,  "zmt_rk":20410826,"zmt_rf":21627457,"swg_rk":588790, "swg_rf":1021132},
    {"outlet":"Parengudi Chennai",   "pos":97078,  "zmt_rk":20486896,"zmt_rf":21087508,"swg_rk":631195, "swg_rf":848486},
  ],
  "Atul Kumar": [
    {"outlet":"Apple Ghar Pune",     "pos":129883, "zmt_rk":20748035,"zmt_rf":21044940,"swg_rk":733937, "swg_rf":741458},
    {"outlet":"Hinjewadi Phase 1",   "pos":141096, "zmt_rk":20855724,"zmt_rf":21049119,"swg_rk":756772, "swg_rf":734618},
    {"outlet":"Millennium Mall Pune","pos":137998, "zmt_rk":21067154,"zmt_rf":None,    "swg_rk":833916, "swg_rf":None},    # RK only
    {"outlet":"Shivaji Nagar Pune",  "pos":346318, "zmt_rk":21435196,"zmt_rf":21604740,"swg_rk":354312, "swg_rf":1009928},
  ],
  "Bhupesh Bhatt": [
    {"outlet":"Madhapur",            "pos":24485,  "zmt_rk":18953624,"zmt_rf":21049126,"swg_rk":120196, "swg_rf":698521},
    {"outlet":"Gachibowli",          "pos":24487,  "zmt_rk":19271816,"zmt_rf":21044950,"swg_rk":214621, "swg_rf":773418},
    {"outlet":"Banjara Hills",       "pos":129436, "zmt_rk":21028217,"zmt_rf":21080628,"swg_rk":711834, "swg_rf":844883},
    {"outlet":"Taranagar Hyderabad", "pos":141099, "zmt_rk":20855101,"zmt_rf":21080636,"swg_rk":766665, "swg_rf":844889},
    {"outlet":"RK Puram Hyderabad",  "pos":44718,  "zmt_rk":19714313,"zmt_rf":None,    "swg_rk":375980, "swg_rf":None},    # RK only
    {"outlet":"Lulu Mall Hyderabad", "pos":141214, "zmt_rk":21154081,"zmt_rf":None,    "swg_rk":866698, "swg_rf":None},    # RK only
    {"outlet":"Miyapur",             "pos":24489,  "zmt_rk":21779883,"zmt_rf":21779942,"swg_rk":1063096,"swg_rf":1061876},
    {"outlet":"Goa Anjuna",          "pos":339817, "zmt_rk":21365117,"zmt_rf":22213849,"swg_rk":946493, "swg_rf":1203077},
  ],
  "Milan": [
    {"outlet":"G Corp",              "pos":367105, "zmt_rk":21734559,"zmt_rf":21865596,"swg_rk":1063584,"swg_rf":1079375},
    {"outlet":"Mumbai Pawai",        "pos":353027, "zmt_rk":21522379,"zmt_rf":21618090,"swg_rk":985771, "swg_rf":1005642},
    {"outlet":"Raymond",             "pos":369670, "zmt_rk":21794492,"zmt_rf":21794546,"swg_rk":1066710,"swg_rf":1066711},
    {"outlet":"Mumbai BKC",          "pos":375018, "zmt_rk":21824216,"zmt_rf":21824173,"swg_rk":1076952,"swg_rf":1102348},
    {"outlet":"Mumbai Chembur",      "pos":383109, "zmt_rk":21966993,"zmt_rf":22030585,"swg_rk":1104174,"swg_rf":1140316},
    {"outlet":"Airoli Navi Mumbai",  "pos":386396, "zmt_rk":21982077,"zmt_rf":22030515,"swg_rk":1123441,"swg_rf":1140313},
    {"outlet":"Mumbai Dahisar",      "pos":399741, "zmt_rk":22170342,"zmt_rf":None,   "swg_rk":1220916,"swg_rf":1276496},  # New Feb 2026
    {"outlet":"Mumbai Marol",        "pos":404829, "zmt_rk":22274879,"zmt_rf":None,   "swg_rk":1238360,"swg_rf":1276481},  # New Feb 2026
    {"outlet":"Mira Road Mumbai",    "pos":386734, "zmt_rk":22044961,"zmt_rf":22150856,"swg_rk":1142360,"swg_rf":1179699},
  ],
}

# ══════════════════════════════════════════════════════════════════════════════
# FILE DETECTION
# ══════════════════════════════════════════════════════════════════════════════
def detect_file_type(file_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheets = [s.lower() for s in wb.sheetnames]
        has_zomato   = any('zomato' in s for s in sheets)
        has_swiggy   = any('swiggy' in s for s in sheets)
        has_foodcost = any('food cost' in s or 'food_cost' in s for s in sheets)
        has_sale     = any('sale' in s for s in sheets)
        if has_zomato and has_swiggy and has_foodcost and has_sale:
            return 'monthly_raw', wb
        return 'unknown', wb
    except:
        return 'unknown', None

# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADERS
# ══════════════════════════════════════════════════════════════════════════════
def load_zomato(wb):
    """Returns dict: {restaurant_id: {orders, complaints, kpt, rating, online_pct}}"""
    zmt = {}
    for sname in wb.sheetnames:
        if 'zomato' not in sname.lower(): continue
        for row in wb[sname].iter_rows(min_row=2, values_only=True):
            rid = safe_id(row[0])
            if not rid: continue
            metric = str(row[5]).strip() if row[5] else ''
            val    = row[6]
            if rid not in zmt:
                zmt[rid] = {'orders':0,'complaints':0,'kpt':None,'rating':None,'online_pct':None}
            if   metric == 'Delivered orders':   zmt[rid]['orders']     = safe_f(val)
            elif metric == 'Total complaints':   zmt[rid]['complaints'] = safe_f(val)
            elif metric == 'KPT (in minutes)':   zmt[rid]['kpt']        = safe_f(val) if val else None
            elif metric == 'Average rating':     zmt[rid]['rating']     = safe_f(val) if val else None
            elif metric == 'Online %':           zmt[rid]['online_pct'] = parse_pct(val)
        break
    return zmt

def load_swiggy(wb):
    """Returns dict: {restaurant_id: {orders, kpt, avail, complaints_count, cmp_pct}}"""
    swg = {}
    for sname in wb.sheetnames:
        if 'swiggy' not in sname.lower(): continue
        for row in wb[sname].iter_rows(min_row=2, values_only=True):
            rid = safe_id(row[0])
            if not rid: continue
            metric = str(row[5]).strip() if row[5] else ''
            val    = row[6]
            if rid not in swg:
                swg[rid] = {'orders':0,'kpt':None,'avail':None,'cmp_pct':None,
                            'missing':0,'quality':0,'quantity':0,'wrong':0,'packaging':0}
            if   metric in ('Delivered Orders','Orders'):
                swg[rid]['orders']    = safe_f(val)
            elif metric == 'Avg Prep Time':
                swg[rid]['kpt']       = parse_min(val)
            elif metric == 'Online Availability %':
                swg[rid]['avail']     = parse_pct(val)
            elif metric == '% Orders with Complaints':
                swg[rid]['cmp_pct']   = parse_pct(val)
            elif metric == 'Missing Items':
                swg[rid]['missing']   = safe_f(val)
            elif metric == 'Quality Issues':
                swg[rid]['quality']   = safe_f(val)
            elif metric == 'Quantity Issues':
                swg[rid]['quantity']  = safe_f(val)
            elif metric == 'Wrong Items':
                swg[rid]['wrong']     = safe_f(val)
            elif metric == 'Packaging & Spillage':
                swg[rid]['packaging'] = safe_f(val)
        break
    return swg

def load_food_cost(wb):
    """
    Returns dict: {pos_id: {fc_pct, net_sale, hygiene_score}}
    Food Cost formula (confirmed):
      FC% = ((Opening Balance + Local/Hyperpure + Store Purchase) - Closing Balance)
             / Net Sale (Net Sale + PC) * 100
    Columns (0-indexed):
      0=Subzone, 1=POS ID, 2=Zone, 3=ASM,
      4=Net Sale + PC (current month),
      5=Opening Balance, 6=Closing Balance,
      7=Local Purchase/Hyperpure, 8=Store Purchase,
      9=Food Cost % (pre-calculated, use if present)
      Last column checked = "Hygiene Score" header
    """
    fc = {}
    for sname in wb.sheetnames:
        if 'food cost' not in sname.lower() and 'food_cost' not in sname.lower(): continue
        ws = wb[sname]

        # Find header row to locate "Hygiene Score" column
        headers = {}
        hyg_col = None
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for ci, val in enumerate(row):
                if val and 'hygiene score' in str(val).lower():
                    hyg_col = ci
                if val and 'hygiene' in str(val).lower() and hyg_col is None:
                    hyg_col = ci
            if hyg_col is not None:
                break

        for row in ws.iter_rows(min_row=2, values_only=True):
            pos = safe_id(row[1])
            if not pos: continue

            net_sale     = safe_f(row[4])
            opening_bal  = safe_f(row[5])
            closing_bal  = safe_f(row[6])
            local_purch  = safe_f(row[7])
            store_purch  = safe_f(row[8])

            # Use pre-calculated FC% if available (col 9)
            fc_raw = row[9] if len(row) > 9 else None
            if fc_raw is not None and str(fc_raw).strip() not in ('','None','nan'):
                fc_pct_val = safe_f(fc_raw)
                # stored as decimal (e.g. 0.38) or percentage (e.g. 38)?
                fc_pct = round(fc_pct_val * 100, 2) if fc_pct_val < 2 else round(fc_pct_val, 2)
            elif net_sale > 0:
                # Calculate: ((Opening + Local + Store) - Closing) / NetSale * 100
                cogs   = (opening_bal + local_purch + store_purch) - closing_bal
                fc_pct = round(cogs / net_sale * 100, 2)
            else:
                fc_pct = None

            # Hygiene Score — read from confirmed column
            hyg = 0
            if hyg_col is not None and hyg_col < len(row):
                hyg = int(safe_f(row[hyg_col]))

            fc[pos] = {'fc_pct': fc_pct, 'net_sale': net_sale, 'hygiene_score': hyg}
        break
    return fc

def load_petpooja_sales(wb):
    """
    Returns dict: {pos_id: net_sale}
    PetPooja Sub-Order Wise report: headers on row 5, data from row 7.
    Outlet name in col A, Net Sales(M.A - T.D) in col G (index 6).
    We match by POS ID cross-referencing the mapping — but PetPooja uses
    outlet names not IDs, so we store by outlet name for MOM comparison.
    """
    sales = {}
    for sname in wb.sheetnames:
        if 'sale' not in sname.lower(): continue
        ws = wb[sname]
        # Find header row
        hdr_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row[0] and 'restaurant' in str(row[0]).lower():
                hdr_row = i; break
            if row[0] and str(row[0]).strip() not in ('','Date:','Name:'):
                if any(v and 'net sale' in str(v).lower() for v in row):
                    hdr_row = i; break
        if not hdr_row: continue
        net_col = None
        for row in ws.iter_rows(min_row=hdr_row, max_row=hdr_row, values_only=True):
            for ci, v in enumerate(row):
                if v and 'net sale' in str(v).lower():
                    net_col = ci; break
        if net_col is None: net_col = 6  # default col G

        current_outlet = None
        for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
            name = str(row[0]).strip() if row[0] else ''
            if not name or name.lower() == 'sub total': continue
            if name.lower() == 'total': continue
            val = safe_f(row[net_col]) if net_col < len(row) else 0
            if val > 0:
                sales[name] = sales.get(name, 0) + val
        break
    return sales

# ══════════════════════════════════════════════════════════════════════════════
# MAIN CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════
def audit_mapping(mapping, zmt, swg, fc, inactive_pos=None):
    """
    Checks every outlet's IDs against what actually exists in the uploaded files.
    Returns a list of discrepancy dicts — one row per missing/mismatched ID.
    """
    inactive_pos = inactive_pos or set()
    issues = []

    zmt_ids = set(zmt.keys())
    swg_ids = set(swg.keys())
    fc_pos  = set(fc.keys())

    for tl, outlets in mapping.items():
        for o in outlets:
            outlet = o['outlet']
            pos    = o['pos']
            if pos in inactive_pos:
                continue

            # ── Zomato RK ─────────────────────────────────────────────────────
            zmt_rk = o.get('zmt_rk')
            if zmt_rk is None:
                issues.append({'tl': tl, 'outlet': outlet, 'platform': 'Zomato',
                                'brand': 'RK', 'id': '—',
                                'issue': 'ID not configured in mapping',
                                'impact': 'Complaints / KPT / Rating / Availability scored 0'})
            elif zmt_rk not in zmt_ids:
                issues.append({'tl': tl, 'outlet': outlet, 'platform': 'Zomato',
                                'brand': 'RK', 'id': str(zmt_rk),
                                'issue': 'ID not found in uploaded Zomato file',
                                'impact': 'Complaints / KPT / Rating / Availability scored 0'})

            # ── Zomato RF (only if RF expected) ───────────────────────────────
            zmt_rf = o.get('zmt_rf')
            if zmt_rf and zmt_rf not in zmt_ids:
                issues.append({'tl': tl, 'outlet': outlet, 'platform': 'Zomato',
                                'brand': 'RF', 'id': str(zmt_rf),
                                'issue': 'ID not found in uploaded Zomato file',
                                'impact': 'RF data excluded from Complaints / Rating / Availability'})

            # ── Swiggy RK ─────────────────────────────────────────────────────
            swg_rk = o.get('swg_rk')
            if swg_rk is None:
                issues.append({'tl': tl, 'outlet': outlet, 'platform': 'Swiggy',
                                'brand': 'RK', 'id': '—',
                                'issue': 'ID not configured in mapping',
                                'impact': 'Complaints / KPT / Availability scored 0 (Swiggy side)'})
            elif swg_rk not in swg_ids:
                issues.append({'tl': tl, 'outlet': outlet, 'platform': 'Swiggy',
                                'brand': 'RK', 'id': str(swg_rk),
                                'issue': 'ID not found in uploaded Swiggy file',
                                'impact': 'Complaints / KPT / Availability scored 0 (Swiggy side)'})

            # ── Swiggy RF (only if RF expected) ───────────────────────────────
            swg_rf = o.get('swg_rf')
            if swg_rf and swg_rf not in swg_ids:
                issues.append({'tl': tl, 'outlet': outlet, 'platform': 'Swiggy',
                                'brand': 'RF', 'id': str(swg_rf),
                                'issue': 'ID not found in uploaded Swiggy file',
                                'impact': 'RF data excluded from Complaints / Availability (Swiggy)'})

            # ── Food Cost / POS ───────────────────────────────────────────────
            if pos and pos not in fc_pos:
                issues.append({'tl': tl, 'outlet': outlet, 'platform': 'Food Cost Sheet',
                                'brand': '—', 'id': str(pos),
                                'issue': 'POS ID not found in uploaded food cost file',
                                'impact': 'Food Cost scored 0 / Hygiene scored 0 / Sale scored 0'})

    return issues


def calculate(mapping, zmt, swg, fc, prev_sales=None, inactive_pos=None):
    results     = []
    disclaimers = []
    flags       = []
    inactive_pos = inactive_pos or set()

    for tl, outlets in mapping.items():
        n           = len([o for o in outlets if o['pos'] not in inactive_pos])
        tl_cmp      = tl_kpt = tl_rat = tl_fc = tl_av = tl_sale = 0
        tl_hyg      = 0
        outlet_rows = []

        for o in outlets:
            outlet = o['outlet']
            pos    = o['pos']
            # Skip outlets marked inactive for this month
            if pos in inactive_pos:
                continue
            # Valid IDs only (None = RF not available for this outlet)
            zids   = [x for x in [o.get('zmt_rk'), o.get('zmt_rf')] if x]
            sids   = [x for x in [o.get('swg_rk'), o.get('swg_rf')] if x]
            rk_only = (o.get('zmt_rf') is None and o.get('swg_rf') is None)
            notes  = []

            # ── COMPLAINTS ────────────────────────────────────────────────────
            # Formula: (Total RK Zomato comps + RK Swiggy comps + RF Zomato comps + RF Swiggy comps)
            #          / (Total orders all platforms) * 100
            # Zomato gives raw complaint counts; Swiggy gives % — back-calc count
            z_orders = sum(zmt[r]['orders']     for r in zids if r in zmt)
            z_comps  = sum(zmt[r]['complaints'] for r in zids if r in zmt)

            s_orders = sum(swg[s]['orders'] for s in sids if s in swg)
            # Back-calculate Swiggy complaint count from % * orders
            s_comps  = 0
            s_has_cmp = False
            for s in sids:
                if s in swg and swg[s].get('cmp_pct') is not None:
                    pct = swg[s]['cmp_pct']
                    ord_ = swg[s].get('orders', 0)
                    s_comps += round(pct / 100 * ord_)
                    s_has_cmp = True

            total_orders = z_orders + s_orders
            total_comps  = z_comps + s_comps

            if total_orders > 0:
                cmp_pct  = round(total_comps / total_orders * 100, 2)
                cmp_pts  = score_complaints(cmp_pct)
                cmp_src  = "RK only" if rk_only else "Swiggy+Zomato"
                if not s_has_cmp and z_orders > 0:
                    cmp_src = "Zomato only (Swiggy missing)"
            else:
                cmp_pct  = None
                cmp_pts  = 0
                cmp_src  = "No data"
                notes.append("No complaint data")
                disclaimers.append(f"{tl} | {outlet}: complaint data missing — scored 0")

            tl_cmp += cmp_pts
            if cmp_pct is not None and cmp_pct >= 3:
                flags.append((tl, outlet, "High Complaints", f"{cmp_pct:.1f}%", ">=3%", cmp_pts))

            # ── KPT ──────────────────────────────────────────────────────────
            # Formula (confirmed): Avg of Zomato RK KPT + Swiggy RK Avg Prep Time
            # RF excluded — confirmed from manual report analysis
            z_rk_id  = o.get('zmt_rk')
            s_rk_id  = o.get('swg_rk')
            kpt_vals = []
            if z_rk_id and z_rk_id in zmt and zmt[z_rk_id].get('kpt') is not None and zmt[z_rk_id]['kpt'] > 0:
                kpt_vals.append(zmt[z_rk_id]['kpt'])
            if s_rk_id and s_rk_id in swg and swg[s_rk_id].get('kpt') is not None and swg[s_rk_id]['kpt'] > 0:
                kpt_vals.append(swg[s_rk_id]['kpt'])

            if kpt_vals:
                avg_kpt = round(sum(kpt_vals) / len(kpt_vals), 2)
                kpt_pts = score_kpt(avg_kpt)
                kpt_src = f"{len(kpt_vals)} readings"
            else:
                avg_kpt = None
                kpt_pts = 0
                kpt_src = "No data"
                notes.append("No KPT data")
                disclaimers.append(f"{tl} | {outlet}: KPT missing — scored 0")

            tl_kpt += kpt_pts
            if avg_kpt is not None and avg_kpt > 0:
                import math
                if math.floor(avg_kpt) > 12:
                    flags.append((tl, outlet, "KPT Exceeded", f"{avg_kpt:.1f} min", "floor>12", kpt_pts))

            # ── RATING ───────────────────────────────────────────────────────
            # Formula: Avg of Average rating [Zomato] for all available IDs
            # If only RK available, use RK only and note it
            rat_vals = [zmt[r]['rating'] for r in zids if r in zmt
                        and zmt[r].get('rating') is not None and zmt[r]['rating'] > 0]
            if rat_vals:
                avg_rat = round(sum(rat_vals) / len(rat_vals), 2)
                rat_pts = score_rating(avg_rat)
                rat_src = "RK only" if rk_only else f"{len(rat_vals)} Zomato IDs"
            else:
                avg_rat = None
                rat_pts = 0
                rat_src = "No data"
                notes.append("No rating data")
                disclaimers.append(f"{tl} | {outlet}: rating missing — scored 0")

            tl_rat += rat_pts
            if avg_rat is not None and avg_rat < 4.0 and avg_rat > 0:
                flags.append((tl, outlet, "Low Rating", f"{avg_rat:.2f}", "<4.0", rat_pts))

            # ── AVAILABILITY ─────────────────────────────────────────────────
            # Formula: Avg of Online % [Zomato] + Online Availability % [Swiggy]
            # for all available IDs (RK + RF where available)
            avail_vals = []
            for r in zids:
                if r in zmt and zmt[r].get('online_pct') is not None:
                    avail_vals.append(zmt[r]['online_pct'])
            for s in sids:
                if s in swg and swg[s].get('avail') is not None:
                    avail_vals.append(swg[s]['avail'])

            if avail_vals:
                avg_avail = round(sum(avail_vals) / len(avail_vals), 2)
                avail_pts = score_avail(avg_avail)
                avail_src = f"{len(avail_vals)} readings"
            else:
                avg_avail = None
                avail_pts = 0
                avail_src = "No data"
                notes.append("No availability data")
                disclaimers.append(f"{tl} | {outlet}: availability missing — scored 0")

            tl_av += avail_pts
            if avg_avail is not None and avg_avail < 98:
                flags.append((tl, outlet, "Low Availability", f"{avg_avail:.1f}%", "<98%", avail_pts))

            # ── FOOD COST ─────────────────────────────────────────────────────
            # Formula (confirmed): ((Opening + Local/Hyperpure + Store) - Closing) / Net Sale * 100
            # Read from food cost sheet by POS ID; Hygiene Score also in same sheet
            fc_data  = fc.get(pos)
            if fc_data and fc_data.get('fc_pct') is not None:
                fc_pct = round(fc_data['fc_pct'], 2)
                fc_pts = score_fc(fc_pct)
                tl_hyg += fc_data.get('hygiene_score', 0)
            else:
                fc_pct = None
                fc_pts = 0
                notes.append("FC data missing")
                disclaimers.append(f"{tl} | {outlet}: food cost missing — scored 0")

            tl_fc += fc_pts
            if fc_pct is not None and fc_pct >= 40:
                flags.append((tl, outlet, "High Food Cost", f"{fc_pct:.1f}%", ">=40%", fc_pts))

            # ── SALE (Month-on-Month via PetPooja Net Sale + PC) ─────────────
            # Scored per outlet: 1pt if this outlet beat last month, 0 if not
            # Summed across all outlets under TL
            curr_ns = fc_data['net_sale'] if fc_data else None
            prev_ns = prev_sales.get(pos, {}).get('net_sale') if prev_sales else None

            if curr_ns is not None and prev_ns is not None and prev_ns > 0:
                outlet_sale_pt = 1 if curr_ns > prev_ns else 0
                sale_src = f"₹{curr_ns:,.0f} vs ₹{prev_ns:,.0f}"
            elif curr_ns is not None and prev_ns is None:
                outlet_sale_pt = 0
                sale_src = "No prev month data"
            else:
                outlet_sale_pt = 0
                sale_src = "No sale data"

            tl_sale += outlet_sale_pt

            outlet_rows.append({
                'outlet':     outlet,
                'pos':        pos,
                'rk_only':    rk_only,
                'cmp_pct':    cmp_pct,   'cmp_pts':   cmp_pts,  'cmp_src':  cmp_src,
                'kpt_avg':    avg_kpt,   'kpt_pts':   kpt_pts,  'kpt_src':  kpt_src,
                'rat_avg':    avg_rat,   'rat_pts':   rat_pts,  'rat_src':  rat_src,
                'avail_pct':  avg_avail, 'avail_pts': avail_pts,'avail_src':avail_src,
                'fc_pct':     fc_pct,    'fc_pts':    fc_pts,
                'hyg_score':  fc_data.get('hygiene_score', 0) if fc_data else 0,
                'curr_sale':  curr_ns,
                'prev_ns':    prev_ns,
                'sale_pts':   outlet_sale_pt,
                'sale_src':   sale_src,
                'notes':      "; ".join(notes) if notes else "OK",
            })

        # ── SALE SUMMARY NOTE ─────────────────────────────────────────────────
        # Sale points already accumulated per-outlet into tl_sale above
        grew_outlets   = sum(1 for r in outlet_rows if r['sale_pts'] == 1)
        total_outlets  = len(outlet_rows)
        curr_tl_sale   = sum(r['curr_sale'] for r in outlet_rows if r['curr_sale'])
        prev_tl_sale   = sum(
            prev_sales.get(o['pos'], {}).get('net_sale', 0)
            for o in outlets
        ) if prev_sales else 0

        if prev_tl_sale > 0:
            sale_note = f"{grew_outlets}/{total_outlets} outlets grew | ₹{curr_tl_sale:,.0f} vs ₹{prev_tl_sale:,.0f}"
        else:
            sale_note = "No prev month data — sale scored 0"

        # ── TOTALS ────────────────────────────────────────────────────────────
        total_pts = tl_cmp + tl_kpt + tl_rat + tl_fc + tl_av + tl_hyg + tl_sale
        avg_score = round(total_pts / n, 2) if n > 0 else 0
        tier      = score_tier(avg_score)

        if prev_tl_sale > 0 and curr_tl_sale < prev_tl_sale:
            flags.append((tl, "ALL OUTLETS", "Sales Decline",
                          f"₹{curr_tl_sale:,.0f}", f"Prev ₹{prev_tl_sale:,.0f}", tl_sale))

        results.append({
            'tl':           tl,
            'outlets':      n,
            'sales_pts':    tl_sale,
            'fc_pts':       tl_fc,
            'cmp_pts':      tl_cmp,
            'kpt_pts':      tl_kpt,
            'rat_pts':      tl_rat,
            'hyg_pts':      tl_hyg,
            'avail_pts':    tl_av,
            'total_pts':    total_pts,
            'avg_score':    avg_score,
            'tier':         tier,
            'sale_note':    sale_note,
            'outlet_detail':outlet_rows,
        })

    return results, disclaimers, flags

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(results, disclaimers, flags, month, discrepancies=None):
    wb = openpyxl.Workbook()

    # ── Light colour palette ──────────────────────────────────────────────────
    HDR   = "2C3E50"
    SUB   = "4A6274"
    WH    = "FFFFFF"
    ALT   = "F5F7FA"
    GN    = "D4EDDA"
    RD    = "FCDEDE"
    YW    = "FFF3CD"
    MG    = "DEE2E6"
    TIER_BG  = {"Platinum":"D4EDDA","Gold":"FFF3CD","Silver":"F0F0F0","Bronze":"FFE8D0"}
    TIER_FG  = {"Platinum":"155724","Gold":"856404","Silver":"495057","Bronze":"6D4C41"}

    def hc(ws, row, col, val, bg=HDR, fg="FFFFFF", bold=True, sz=9, align="center", wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, color=fg, size=sz, name="Arial")
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border    = thin_border()
        return c

    def dc(ws, row, col, val, bg=WH, fg="222222", bold=False, align="center", fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, color=fg, size=9, name="Arial")
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        if fmt: c.number_format = fmt
        c.border    = thin_border()
        return c

    sorted_r = sorted(results, key=lambda x: x['avg_score'], reverse=True)

    # ── SHEET 1: TL Performance Summary ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "TL Performance Summary"
    ws1.freeze_panes = "A4"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:H1")
    c = ws1["A1"]
    c.value = f"RollsKing — Monthly Performance Report | {month}"
    c.font  = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    c.fill  = PatternFill("solid", start_color=HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:H2")
    c = ws1["A2"]
    c.value = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}  |  {len(results)} Team Leaders  |  {sum(r['outlets'] for r in results)} Outlets"
    c.font  = Font(size=8, color="AAAAAA", name="Arial")
    c.fill  = PatternFill("solid", start_color=SUB)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 16

    # Headers — Sale Pts + Hygiene Pts only; NO FC/Cmp/KPT/Rating/Avail pts
    hdrs = [("Team Leader",26),("Outlets",9),("Sale Pts",9),("Hygiene Pts",11),
            ("Total Pts",10),("Avg Score",10),("Tier",12)]
    ws1.row_dimensions[3].height = 24
    for ci,(h,w) in enumerate(hdrs,1):
        hc(ws1, 3, ci, h)
        ws1.column_dimensions[get_column_letter(ci)].width = w

    for ri, r in enumerate(sorted_r, 4):
        bg = ALT if ri % 2 == 0 else WH
        tbg = TIER_BG.get(r['tier'], WH)
        tfg = TIER_FG.get(r['tier'], "222222")
        vals = [r['tl'], r['outlets'], r['sales_pts'], r['hyg_pts'],
                r['total_pts'], r['avg_score'], r['tier']]
        for ci, v in enumerate(vals, 1):
            cell_bg = tbg if ci == 7 else (WH if ci == 1 else bg)
            cell_fg = tfg if ci == 7 else ("222222")
            dc(ws1, ri, ci, v, bg=cell_bg, fg=cell_fg,
               bold=(ci in (1,7)), align="left" if ci == 1 else "center")
        ws1.row_dimensions[ri].height = 18

    # Grand total row
    tr = len(sorted_r) + 4
    for ci in range(1, 8):
        c = ws1.cell(row=tr, column=ci)
        c.font  = Font(bold=True, size=9, color="FFFFFF", name="Arial")
        c.fill  = PatternFill("solid", start_color=HDR)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws1["A"+str(tr)].value = "GRAND TOTAL"
    ws1["A"+str(tr)].alignment = Alignment(horizontal="left", vertical="center")
    ws1.cell(row=tr,column=3).value = sum(r['sales_pts'] for r in results)
    ws1.cell(row=tr,column=4).value = sum(r['hyg_pts']   for r in results)
    ws1.cell(row=tr,column=5).value = sum(r['total_pts'] for r in results)
    ws1.row_dimensions[tr].height = 20

    # Scoring reference
    sr = tr + 2
    ws1.cell(row=sr, column=1, value="SCORING REFERENCE").font = Font(bold=True, size=9, color=HDR, name="Arial")
    conds = [
        ("Sale",         "1pt per outlet that beat prev month Net Sale — summed per TL",    "PetPooja food cost sheet"),
        ("Food Cost",    "< 40% = 1pt | >= 40% = 0pt",                                      "Food cost sheet (calculated)"),
        ("Complaints",   "0-<1% = 4pts | 1-<2% = 3pts | 2-<3% = 1pt | >=3% = 0pt",        "Zomato + Swiggy blended"),
        ("KPT",          "floor(avg) <= 12min = 1pt | >12min = 0pt",                        "Avg of Zomato RK KPT + Swiggy RK Avg Prep Time"),
        ("Rating",       ">= 4.0 = 1pt | < 4.0 = 0pt",                                     "Zomato Average Rating (all available IDs)"),
        ("Hygiene",      "Sum of outlet hygiene scores (manually added to food cost sheet)", "Hygiene Score column in food cost file"),
        ("Availability", ">= 98% = 1pt | < 98% = 0pt",                                     "Avg of Zomato Online% + Swiggy Online Availability%"),
        ("Grade",        "Bronze 0-3 | Silver 3-6 | Gold 6-8 | Platinum 8+",               "Total Pts / No. of outlets"),
    ]
    for j, (k, rule, src) in enumerate(conds, sr+1):
        bg = ALT if j % 2 == 0 else WH
        for ci, v in enumerate([k, rule, src], 1):
            c = ws1.cell(row=j, column=ci, value=v)
            c.font = Font(size=8, bold=(ci==1), name="Arial", color="222222")
            c.fill = PatternFill("solid", start_color=bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws1.row_dimensions[j].height = 18
    ws1.merge_cells(f"B{sr+1}:C{sr+1}")
    for j in range(sr+1, sr+len(conds)+1):
        ws1.merge_cells(f"B{j}:C{j}")

    # ── SHEET 2: Outlet Detail ────────────────────────────────────────────────
    # Columns: TL | Outlet | Brands | Net Sale | Prev Sale | FC% | Cmp% | KPT | Rating | Avail% | Hygiene | Notes
    ws2 = wb.create_sheet("Outlet Detail")
    ws2.freeze_panes = "C3"
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:L1")
    c = ws2["A1"]
    c.value = f"Outlet Detail — {month}"
    c.font  = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    c.fill  = PatternFill("solid", start_color=HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    od_hdrs = [
        ("Team Leader",20), ("Outlet",24), ("Brands",10),
        ("Net Sale (₹)",14), ("Prev Sale (₹)",14),
        ("Food Cost %",11), ("Complaint %",11),
        ("KPT (min)",10), ("Rating",9), ("Availability %",13),
        ("Hygiene",9), ("Notes",28),
    ]
    for ci,(h,w) in enumerate(od_hdrs,1):
        hc(ws2, 2, ci, h, bg=SUB, wrap=True)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 28

    row_num = 3
    for r in sorted_r:
        for o in r['outlet_detail']:
            bg = ALT if row_num % 2 == 0 else WH
            brand = "RK" if o['rk_only'] else "RK + RF"
            fc_bg  = RD if (o['fc_pct'] and o['fc_pct'] >= 40) else (GN if o['fc_pct'] else bg)
            cmp_bg = RD if (o['cmp_pct'] and o['cmp_pct'] >= 3) else (YW if (o['cmp_pct'] and o['cmp_pct'] >= 2) else (GN if o['cmp_pct'] is not None else bg))
            kpt_bg = RD if (o['kpt_avg'] and int(o['kpt_avg']) > 12) else (GN if o['kpt_avg'] else bg)
            rat_bg = RD if (o['rat_avg'] and o['rat_avg'] < 4.0) else (GN if o['rat_avg'] else bg)
            av_bg  = RD if (o['avail_pct'] and o['avail_pct'] < 98) else (GN if o['avail_pct'] else bg)

            vals = [
                r['tl'], o['outlet'], brand,
                o.get('curr_sale'), o.get('prev_ns'),   # net sale, prev sale
                o['fc_pct'], o['cmp_pct'],
                o['kpt_avg'], o['rat_avg'], o['avail_pct'],
                o['hyg_score'], o['notes'],
            ]
            bgs = [WH, bg, bg, bg, bg, fc_bg, cmp_bg, kpt_bg, rat_bg, av_bg, bg, bg]
            aligns = ["left","left","center","center","center","center","center","center","center","center","center","left"]
            fmts   = [None,None,None,'#,##0','#,##0','0.00','0.00','0.0','0.00','0.0',None,None]
            for ci,(v,b,al,fm) in enumerate(zip(vals,bgs,aligns,fmts),1):
                dc(ws2, row_num, ci, v, bg=b, align=al, fmt=fm, bold=(ci==1))
            ws2.row_dimensions[row_num].height = 16
            row_num += 1

    # ── SHEET 3: Flagged Outlets ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Flagged Outlets")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:E1")
    c = ws3["A1"]
    c.value = f"Flagged Outlets — {month}  ({len(flags)} issues)"
    c.font  = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    c.fill  = PatternFill("solid", start_color="8B0000")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22
    for ci,(h,w) in enumerate([("Team Leader",20),("Outlet",24),("Issue",22),("Value",14),("Threshold",14)],1):
        hc(ws3, 2, ci, h, bg="8B0000")
        ws3.column_dimensions[get_column_letter(ci)].width = w
    flag_bg = {"High Complaints":RD,"KPT Exceeded":YW,"Low Rating":YW,
               "High Food Cost":RD,"Low Availability":YW,"Sales Decline":RD}
    for i,(tl,outlet,issue,val,thresh,pts) in enumerate(flags,3):
        bg = flag_bg.get(issue, WH)
        for ci,v in enumerate([tl,outlet,issue,val,thresh],1):
            dc(ws3, i, ci, v, bg=bg, align="left" if ci<=2 else "center")
        ws3.row_dimensions[i].height = 15

    # ── SHEET 4: Data Discrepancy ─────────────────────────────────────────────
    discrepancies = discrepancies or []
    ws4 = wb.create_sheet("Data Discrepancy")
    ws4.sheet_view.showGridLines = False

    DISC_HDR = "7B2D00"  # dark orange — distinct from flags red
    disc_cols = [
        ("Team Leader", 20), ("Outlet", 24), ("Platform", 16),
        ("Brand", 8), ("ID in Mapping", 16),
        ("Issue", 42), ("Impact on Score", 48),
    ]

    ws4.merge_cells(f"A1:{get_column_letter(len(disc_cols))}1")
    c = ws4["A1"]
    disc_count = len(discrepancies)
    c.value = (
        f"Data Discrepancy Report — {month}  |  "
        f"{disc_count} ID issue{'s' if disc_count != 1 else ''} detected"
        if disc_count > 0
        else f"Data Discrepancy Report — {month}  |  ✅ All IDs matched — no issues found"
    )
    c.font  = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    c.fill  = PatternFill("solid", start_color=DISC_HDR if disc_count > 0 else "1A5C2A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 26

    # Sub-header explaining the sheet
    ws4.merge_cells(f"A2:{get_column_letter(len(disc_cols))}2")
    c = ws4["A2"]
    c.value = (
        "These IDs are in the app mapping but were NOT found in the uploaded files. "
        "The report still runs — affected metrics are scored 0. "
        "Fix by updating the outlet mapping with the correct IDs."
        if disc_count > 0
        else "Every outlet ID in the mapping was found in the uploaded Zomato, Swiggy, and Food Cost files."
    )
    c.font = Font(size=8, color="FFFFFF", italic=True, name="Arial")
    c.fill = PatternFill("solid", start_color="B34700" if disc_count > 0 else "2D7A3A")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.row_dimensions[2].height = 28

    for ci, (h, w) in enumerate(disc_cols, 1):
        hc(ws4, 3, ci, h, bg=DISC_HDR if disc_count > 0 else "1A5C2A")
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[3].height = 20

    if discrepancies:
        # Group by TL for visual clarity
        current_tl = None
        row_num4 = 4
        for d in sorted(discrepancies, key=lambda x: (x['tl'], x['outlet'], x['platform'])):
            # TL separator row
            if d['tl'] != current_tl:
                current_tl = d['tl']
                ws4.merge_cells(f"A{row_num4}:{get_column_letter(len(disc_cols))}{row_num4}")
                c = ws4.cell(row=row_num4, column=1, value=f"  {current_tl}")
                c.font  = Font(bold=True, size=9, color="FFFFFF", name="Arial")
                c.fill  = PatternFill("solid", start_color="4A3000")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = thin_border()
                ws4.row_dimensions[row_num4].height = 16
                row_num4 += 1

            bg = "FFF0E6"  # light orange tint for all discrepancy rows
            vals = [
                d['tl'], d['outlet'], d['platform'],
                d['brand'], d['id'], d['issue'], d['impact'],
            ]
            aligns = ["left", "left", "center", "center", "center", "left", "left"]
            for ci, (v, al) in enumerate(zip(vals, aligns), 1):
                cell = ws4.cell(row=row_num4, column=ci, value=v)
                cell.font = Font(
                    size=8, name="Arial",
                    bold=(ci in (1, 2)),
                    color="5C1A00" if ci >= 6 else "222222"
                )
                cell.fill = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(horizontal=al, vertical="center", wrap_text=(ci >= 6))
                cell.border = thin_border()
            ws4.row_dimensions[row_num4].height = 28
            row_num4 += 1
    else:
        # All clear row
        ws4.merge_cells(f"A4:{get_column_letter(len(disc_cols))}4")
        c = ws4.cell(row=4, column=1,
                     value="✅  All outlet IDs matched successfully. No discrepancies found.")
        c.font = Font(bold=True, size=10, color="155724", name="Arial")
        c.fill = PatternFill("solid", start_color="D4EDDA")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws4.row_dimensions[4].height = 24

    # ── SHEET 5: Data Notes ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Data Notes")
    ws5.merge_cells("A1:B1")
    c = ws5["A1"]
    c.value = f"Data Notes — {month}"
    c.font  = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    c.fill  = PatternFill("solid", start_color=HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 22
    for ci,(h,w) in enumerate([("Outlet / TL",38),("Note",55)],1):
        hc(ws5, 2, ci, h, bg=SUB)
        ws5.column_dimensions[get_column_letter(ci)].width = w
    for i,note in enumerate(disclaimers,3):
        parts = note.split("|",1)
        for ci,v in enumerate([parts[0].strip(), parts[1].strip() if len(parts)>1 else ""],1):
            c = ws5.cell(row=i, column=ci, value=v)
            c.font = Font(size=8, name="Arial", color="444444")
            c.border = thin_border()
        ws5.row_dimensions[i].height = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(results, flags, disclaimers, month):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    plt.close('all')
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable, Image as RLImage)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    DARK   = colors.HexColor("#2C3E50")
    LIGHT  = colors.HexColor("#F8F9FA")
    GOLD   = colors.HexColor("#F39C12")
    SILV   = colors.HexColor("#95A5A6")
    BRNZ   = colors.HexColor("#CD7F32")
    PLAT   = colors.HexColor("#27AE60")
    TIER_C = {"Platinum":PLAT,"Gold":GOLD,"Silver":SILV,"Bronze":BRNZ}
    TIER_TXT = {"Platinum":colors.white,"Gold":colors.white,
                "Silver":colors.HexColor("#222222"),"Bronze":colors.white}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=12*mm, bottomMargin=12*mm,
                            leftMargin=14*mm, rightMargin=14*mm)
    styles = getSampleStyleSheet()

    title_sty = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=16,
                                textColor=DARK, alignment=TA_CENTER, spaceAfter=3)
    sub_sty   = ParagraphStyle("s", fontName="Helvetica", fontSize=9,
                                textColor=colors.HexColor("#888888"),
                                alignment=TA_CENTER, spaceAfter=10)
    sec_sty   = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=11,
                                textColor=DARK, spaceBefore=8, spaceAfter=5)

    sorted_r = sorted(results, key=lambda x: x['avg_score'], reverse=True)
    story = []

    # ── TITLE ──
    story.append(Paragraph("RollsKing — Monthly Performance Report", title_sty))
    story.append(Paragraph(
        f"{month}  |  {datetime.now().strftime('%d %b %Y, %H:%M')}  |  "
        f"{len(results)} Team Leaders  |  {sum(r['outlets'] for r in results)} Outlets",
        sub_sty))
    story.append(HRFlowable(width="100%", thickness=1, color=DARK, spaceAfter=8))

    # ── SECTION 1: PERFORMANCE SUMMARY TABLE ──────────────────────────────────
    story.append(Paragraph("Performance Summary", sec_sty))

    tbl_data = [["Team Leader","Outlets","Sale","FC","Cmp","KPT","Rating","Hygiene","Avail","Total","Avg","Tier"]]
    tier_idx = {}
    for ri, r in enumerate(sorted_r, 1):
        tbl_data.append([
            r['tl'], r['outlets'],
            r['sales_pts'], r['fc_pts'], r['cmp_pts'],
            r['kpt_pts'], r['rat_pts'], r['hyg_pts'], r['avail_pts'],
            r['total_pts'], f"{r['avg_score']:.2f}", r['tier']
        ])
        tier_idx[ri] = r['tier']

    cw = [55*mm,13*mm,12*mm,10*mm,11*mm,11*mm,13*mm,13*mm,11*mm,13*mm,12*mm,16*mm]
    ts = [
        ("BACKGROUND",  (0,0),(-1,0), DARK),
        ("TEXTCOLOR",   (0,0),(-1,0), colors.white),
        ("FONTNAME",    (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0),(-1,-1), 8),
        ("ALIGN",       (0,0),(-1,-1), "CENTER"),
        ("ALIGN",       (0,1),(0,-1),  "LEFT"),
        ("FONTNAME",    (0,1),(0,-1),  "Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F5F7FA")]),
        ("GRID",        (0,0),(-1,-1), 0.4, colors.HexColor("#DDDDDD")),
        ("TOPPADDING",  (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING", (0,0),(-1,-1), 4),
    ]
    for ri, tier in tier_idx.items():
        tc = TIER_C.get(tier, SILV)
        tt = TIER_TXT.get(tier, colors.white)
        ts += [("BACKGROUND",(11,ri),(11,ri),tc),
               ("TEXTCOLOR", (11,ri),(11,ri),tt),
               ("FONTNAME",  (11,ri),(11,ri),"Helvetica-Bold")]

    tbl = Table(tbl_data, colWidths=cw)
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)
    story.append(Spacer(1, 10*mm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor("#DDDDDD"), spaceAfter=8))

    # ── SECTION 2: TIER DISTRIBUTION PIE CHART ────────────────────────────────
    story.append(Paragraph("Grade Distribution", sec_sty))

    tier_counts = {"Platinum":0,"Gold":0,"Silver":0,"Bronze":0}
    for r in results:
        tier_counts[r["tier"]] = tier_counts.get(r["tier"],0) + 1
    active = {k:v for k,v in tier_counts.items() if v > 0}
    pie_clr = {"Platinum":"#27AE60","Gold":"#F39C12","Silver":"#95A5A6","Bronze":"#CD7F32"}

    fig, ax = plt.subplots(figsize=(5, 3.5), facecolor="white")
    wedges, texts, autotexts = ax.pie(
        active.values(),
        labels=[f"{k} ({v})" for k,v in active.items()],
        colors=[pie_clr[k] for k in active],
        autopct="%1.0f%%",
        startangle=140,
        pctdistance=0.72,
        textprops={"fontsize":9,"color":"#2C3E50"},
        wedgeprops={"edgecolor":"white","linewidth":2},
    )
    for at in autotexts:
        at.set_color("white"); at.set_fontweight("bold"); at.set_fontsize(9)
    ax.set_title("TL Tier Breakdown", fontsize=10, color="#2C3E50",
                 fontweight="bold", pad=8)
    plt.tight_layout()

    img_buf = io.BytesIO()
    fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    img_buf.seek(0)
    story.append(RLImage(img_buf, width=110*mm, height=78*mm))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for k, v in [('logged_in',False),('report_bytes',None),('report_name',None),
              ('pdf_bytes',None),('pdf_name',None)]:
    if k not in st.session_state: st.session_state[k] = v
if 'mapping' not in st.session_state:
    st.session_state.mapping = {k: list(v) for k, v in DEFAULT_MAPPING.items()}
if 'inactive_pos' not in st.session_state:
    st.session_state.inactive_pos = set()  # set of POS IDs marked inactive this session

# ══════════════════════════════════════════════════════════════════════════════
# LOGIN
# ══════════════════════════════════════════════════════════════════════════════
APP_PASSWORD = "rollsking2025"
if not st.session_state.logged_in:
    st.markdown('<div class="main-title">RollsKing</div>', unsafe_allow_html=True)
    st.markdown('<div class="main-subtitle">Operations Report Generator</div>', unsafe_allow_html=True)
    pw = st.text_input("Password", type="password", placeholder="Enter access password")
    if st.button("Sign In"):
        if pw == APP_PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="main-title">RollsKing Reports</div>', unsafe_allow_html=True)
st.markdown('<div class="main-subtitle">Monthly Performance Report Generator</div>', unsafe_allow_html=True)

mapping       = st.session_state.mapping
tl_names      = sorted(mapping.keys())
total_outlets = sum(len(v) for v in mapping.values())

tab_report, tab_mapping, tab_outlets = st.tabs(["📊  Generate Report", "🗂️  Manage Mapping", "🏪  Outlet Status"])

# ── TAB 1: GENERATE REPORT ────────────────────────────────────────────────────
with tab_report:
    st.markdown(f"""
    <div class="card">
        <div class="section-label">Network Status</div>
        <span class="status-ok">✓ {len(tl_names)} Team Leaders &nbsp;·&nbsp; {total_outlets} Outlets</span>
        <div style="color:#555;font-size:0.78rem;margin-top:0.3rem;">
            Mapping hardcoded from confirmed master sheet. IDs used for matching — name variations ignored.
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Step 1 — Upload
    st.markdown("""<div style="margin-bottom:0.5rem;">
        <span class="step-badge">1</span>
        <span style="font-family:'Syne',sans-serif;font-size:0.9rem;font-weight:700;color:#fff;">Upload Data Files</span>
        <span style="color:#555;font-size:0.8rem;margin-left:8px;">Upload current month + previous month for Sale scoring</span>
    </div>""", unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Upload .xlsx files", type=["xlsx"],
        accept_multiple_files=True, label_visibility="collapsed")

    detected = []
    if uploaded_files:
        for f in uploaded_files:
            fbytes = f.read()
            ftype, wb = detect_file_type(fbytes)
            detected.append({"name":f.name,"type":ftype,"bytes":fbytes,"wb":wb})
        for d in detected:
            ok  = d["type"] == "monthly_raw"
            css = "chip-ok" if ok else "chip-warn"
            lbl = "Monthly Data ✓" if ok else "⚠ Needs: Zomato + Swiggy + Food Cost + Sale sheets"
            st.markdown(f'<span class="{css}">{"✓" if ok else "⚠"} {d["name"]} — {lbl}</span>',
                        unsafe_allow_html=True)

    st.markdown("<div style='margin:1rem 0;'></div>", unsafe_allow_html=True)

    # Step 2 — Month
    st.markdown("""<div style="margin-bottom:0.5rem;">
        <span class="step-badge">2</span>
        <span style="font-family:'Syne',sans-serif;font-size:0.9rem;font-weight:700;color:#fff;">Select Report Month</span>
    </div>""", unsafe_allow_html=True)
    months    = ["January 2026","February 2026","March 2026","April 2026","May 2026","June 2026",
                 "July 2026","August 2026","September 2026","October 2026","November 2025","December 2025"]
    sel_month = st.selectbox("Month", months, label_visibility="collapsed")

    st.markdown("<div style='margin:1rem 0;'></div>", unsafe_allow_html=True)

    # Step 3 — Generate
    st.markdown("""<div style="margin-bottom:0.5rem;">
        <span class="step-badge">3</span>
        <span style="font-family:'Syne',sans-serif;font-size:0.9rem;font-weight:700;color:#fff;">Generate Report</span>
        <span style="color:#555;font-size:0.8rem;margin-left:8px;">Hygiene scores are read automatically from the food cost sheet</span>
    </div>""", unsafe_allow_html=True)

    valid = [d for d in detected if d["type"] == "monthly_raw"] if detected else []

    if not valid:
        st.markdown("""<div style="background:#1a1a1a;border:1px dashed #333;border-radius:10px;
        padding:1rem;color:#555;font-size:0.85rem;text-align:center;">
            Upload monthly data file above to enable report generation
        </div>""", unsafe_allow_html=True)
    else:
        if st.button("⚡ Generate Report"):
            with st.spinner("Processing data..."):
                try:
                    curr_key   = sel_month[:3].lower()
                    curr_files = [d for d in valid if curr_key in d["name"].lower()]
                    prev_files = [d for d in valid if curr_key not in d["name"].lower()]
                    if not curr_files: curr_files = valid

                    # Load current month
                    all_zmt, all_swg, all_fc = {}, {}, {}
                    for d in curr_files:
                        all_zmt.update(load_zomato(d["wb"]))
                        all_swg.update(load_swiggy(d["wb"]))
                        all_fc.update(load_food_cost(d["wb"]))

                    # Load previous month food cost for sale comparison
                    prev_fc = {}
                    for d in prev_files:
                        prev_fc.update(load_food_cost(d["wb"]))

                    # ── Run ID audit BEFORE calculating ──────────────────────
                    discrepancies = audit_mapping(
                        mapping, all_zmt, all_swg, all_fc,
                        inactive_pos=st.session_state.inactive_pos
                    )

                    results, disclaimers, flags = calculate(
                        mapping, all_zmt, all_swg, all_fc,
                        prev_sales=prev_fc if prev_fc else None,
                        inactive_pos=st.session_state.inactive_pos
                    )

                    month_slug  = sel_month.replace(" ","_")
                    excel_bytes = build_excel(results, disclaimers, flags, sel_month,
                                             discrepancies=discrepancies)
                    st.session_state.report_bytes = excel_bytes
                    st.session_state.report_name  = f"RollsKing_Report_{month_slug}.xlsx"

                    try:
                        pdf_bytes = build_pdf(results, flags, disclaimers, sel_month)
                        st.session_state.pdf_bytes = pdf_bytes
                        st.session_state.pdf_name  = f"RollsKing_Report_{month_slug}.pdf"
                        pdf_ok = True
                    except Exception as pdf_err:
                        import traceback
                        st.session_state.pdf_bytes = None
                        st.session_state._pdf_error = traceback.format_exc()
                        pdf_ok = False

                    total_sale_pts = sum(r['sales_pts'] for r in results)
                    total_possible = sum(r['outlets'] for r in results)
                    sale_note = f" · {total_sale_pts}/{total_possible} outlets grew sales" if prev_fc \
                                else " · Upload prev month file for sale scoring"

                    st.success(f"✓ Report ready — {len(results)} TLs · "
                               f"{sum(r['outlets'] for r in results)} Outlets · "
                               f"{len(flags)} Flags{sale_note}")

                    # Show discrepancy warning inline if issues found
                    if discrepancies:
                        tls_affected = len(set(d['tl'] for d in discrepancies))
                        st.warning(
                            f"⚠️ **{len(discrepancies)} ID issue{'s' if len(discrepancies)!=1 else ''} detected** "
                            f"across {tls_affected} TL{'s' if tls_affected!=1 else ''} — "
                            f"affected metrics scored 0. See **Data Discrepancy** sheet in the Excel download for full details."
                        )
                    else:
                        st.info("✅ All outlet IDs matched — no data discrepancies found.")

                    if not pdf_ok:
                        st.warning("PDF failed — see error below:")
                        st.code(st.session_state.get('_pdf_error','unknown error'))

                except Exception as e:
                    st.error(f"Error: {e}")
                    import traceback
                    st.code(traceback.format_exc())

    # Downloads
    if st.session_state.report_bytes:
        st.markdown("<hr class='divider'>", unsafe_allow_html=True)
        st.markdown('<div class="section-label">Download Reports</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            st.download_button("📥 Download Excel",
                data=st.session_state.report_bytes,
                file_name=st.session_state.report_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with c2:
            if st.session_state.pdf_bytes:
                st.download_button("📄 Download PDF",
                    data=st.session_state.pdf_bytes,
                    file_name=st.session_state.pdf_name,
                    mime="application/pdf")
            else:
                err = st.session_state.get('_pdf_error','')
                if err:
                    with st.expander("⚠ PDF error — click to see"):
                        st.code(err)

# ── TAB 2: MANAGE MAPPING ─────────────────────────────────────────────────────
with tab_mapping:
    st.markdown(f"""
    <div class="card">
        <div class="section-label">Current Mapping</div>
        <span class="status-ok">✓ {len(tl_names)} Team Leaders · {total_outlets} Outlets</span>
        <div style="color:#555;font-size:0.78rem;margin-top:0.3rem;">
            Hardcoded from confirmed master sheet. Changes here are session-only.
        </div>
    </div>
    """, unsafe_allow_html=True)

    for tl, outlets in sorted(mapping.items()):
        with st.expander(f"{tl}  —  {len(outlets)} outlets"):
            for o in outlets:
                rk_tag = " 🔵 RK only" if (o.get('zmt_rf') is None and o.get('swg_rf') is None) else ""
                st.markdown(
                    f"**{o['outlet']}**{rk_tag} &nbsp;·&nbsp; POS: `{o['pos']}` "
                    f"&nbsp;·&nbsp; Z_RK: `{o.get('zmt_rk')}` Z_RF: `{o.get('zmt_rf')}` "
                    f"&nbsp;·&nbsp; S_RK: `{o.get('swg_rk')}` S_RF: `{o.get('swg_rf')}`"
                )

    st.markdown("<div style='margin:1.2rem 0;'></div>", unsafe_allow_html=True)

    st.markdown("""<div style="margin-bottom:0.4rem;">
        <span class="step-badge">+</span>
        <span style="font-family:'Syne',sans-serif;font-size:0.85rem;font-weight:700;color:#fff;">Add New Outlet</span>
    </div>""", unsafe_allow_html=True)
    sel_tl = st.selectbox("Assign to Team Leader", sorted(mapping.keys()), key="add_tl")
    c1, c2 = st.columns(2)
    with c1:
        new_outlet = st.text_input("Outlet Name",               placeholder="e.g. Sector 62 Noida", key="new_outlet")
        new_pos    = st.text_input("POS ID (required)",         placeholder="e.g. 23687",           key="new_pos")
        new_zrk    = st.text_input("Zomato ID (RollsKing)",     placeholder="e.g. 19476740",        key="new_zrk")
        new_zrf    = st.text_input("Zomato ID (Rolling Fresh)", placeholder="leave blank if RK only",key="new_zrf")
    with c2:
        new_srk    = st.text_input("Swiggy ID (RollsKing)",     placeholder="e.g. 313666",          key="new_srk")
        new_srf    = st.text_input("Swiggy ID (Rolling Fresh)", placeholder="leave blank if RK only",key="new_srf")

    if st.button("➕ Add Outlet"):
        if new_outlet and new_pos:
            st.session_state.mapping[sel_tl].append({
                "outlet": new_outlet.strip(),
                "pos":    safe_id(new_pos),
                "zmt_rk": safe_id(new_zrk) if new_zrk else None,
                "zmt_rf": safe_id(new_zrf) if new_zrf else None,
                "swg_rk": safe_id(new_srk) if new_srk else None,
                "swg_rf": safe_id(new_srf) if new_srf else None,
            })
            st.success(f"✓ {new_outlet} added under {sel_tl}")
            st.rerun()
        else:
            st.warning("Outlet Name and POS ID are required.")

# ── TAB 3: OUTLET STATUS ──────────────────────────────────────────────────────
with tab_outlets:
    st.markdown(f"""
    <div class="card">
        <div class="section-label">Outlet Status — This Session</div>
        <div style="color:#aaa;font-size:0.85rem;">
            Mark outlets as <b style="color:#ef4444;">Inactive</b> to exclude them from this month's report.
            Useful when an outlet is temporarily closed or has no data for the month.
            <br><br>
            <span style="color:#555;font-size:0.78rem;">Changes are session-only and reset when you refresh the page.</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    inactive = st.session_state.inactive_pos

    # Show all TLs with their outlets as toggles
    for tl, outlets in sorted(mapping.items()):
        active_count   = sum(1 for o in outlets if o['pos'] not in inactive)
        inactive_count = sum(1 for o in outlets if o['pos'] in inactive)
        badge = f"<span style='color:#4ade80;font-size:0.78rem;'>{active_count} active</span>"
        if inactive_count:
            badge += f" &nbsp; <span style='color:#ef4444;font-size:0.78rem;'>{inactive_count} inactive</span>"

        with st.expander(f"{tl}  —  {len(outlets)} outlets  |  {active_count} active"):
            cols = st.columns(2)
            for i, o in enumerate(outlets):
                pos    = o['pos']
                is_off = pos in inactive
                with cols[i % 2]:
                    label = f"{'🔴' if is_off else '🟢'}  {o['outlet']}"
                    toggle = st.checkbox(
                        label,
                        value=(not is_off),
                        key=f"status_{pos}"
                    )
                    if not toggle and pos not in inactive:
                        st.session_state.inactive_pos.add(pos)
                        st.rerun()
                    elif toggle and pos in inactive:
                        st.session_state.inactive_pos.discard(pos)
                        st.rerun()

    st.markdown("<div style='margin:1rem 0;'></div>", unsafe_allow_html=True)

    # Summary
    total_active   = sum(len(v) for v in mapping.values()) - len(inactive)
    total_inactive = len(inactive)
    if total_inactive > 0:
        st.markdown(f"""
        <div class="card">
            <div class="section-label">Currently Inactive ({total_inactive} outlets)</div>
            <div style="color:#ef4444;font-size:0.85rem;">
            These outlets will be excluded from the next report generated.
            </div>
        </div>
        """, unsafe_allow_html=True)
        for tl, outlets in sorted(mapping.items()):
            for o in outlets:
                if o['pos'] in inactive:
                    st.markdown(f"&nbsp;&nbsp;🔴 **{o['outlet']}** — {tl} (POS: {o['pos']})")

        if st.button("🔄 Reset — Mark All Outlets Active"):
            st.session_state.inactive_pos = set()
            st.rerun()
    else:
        st.markdown("""
        <div style="background:#1a1a1a;border:1px solid #2a2a2a;border-radius:10px;
        padding:1rem;color:#4ade80;font-size:0.85rem;text-align:center;">
            ✓ All outlets active — no exclusions for this report
        </div>
        """, unsafe_allow_html=True)

st.markdown("""
<div style='text-align:center;color:#2a2a2a;font-size:0.75rem;padding:2rem 0 1rem;'>
    RollsKing Internal Tools · Built for Operations
</div>
""", unsafe_allow_html=True)
